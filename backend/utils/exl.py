from pathlib import Path
from typing import List

import openpyxl
from django.conf import settings
from openpyxl.utils import get_column_letter


class Exl:
    """
    Класс для создания и сохранения отчетов в формате Excel (.xlsx).

    @method create: Создает xlsx файл на основе переданных данных и сохраняет его по указанному пути.
    """

    @staticmethod
    def create(data: List[List], path: str = settings.REPORTS_PATH) -> str:
        """
        Генерирует и сохраняет Excel файл на основе переданных данных.

        @param data: Двумерный список, где каждая вложенная коллекция — строка таблицы.
        @param path: Путь, куда сохранить сгенерированный файл. По умолчанию используется `settings.REPORTS_PATH`.
        @return: Возвращает полный путь к сохраненному файлу.
        @raise ValueError: Если данные не представлены в виде списка списков.
        """
        if not all(isinstance(row, list) for row in data):
            raise ValueError("Data must be a list of lists.")

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                col_letter = get_column_letter(col_idx)
                sheet[f'{col_letter}{row_idx}'] = value

        Path(path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path
